import xlrd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import sys
from os import path
from os import mkdir
import settings
import utils
from settings import IniEnum
from datetime import datetime


class SetupFilePath(object):
    def __init__(self):
        super(SetupFilePath, self).__init__()
        self.root_path = sys.path[0]
        self.result_path = None
        self.setting_ini_path = None
        self.template_path = None

        self.find_root_path()
        self.set_all_path()

    def find_root_path(self):
        root_path_find = False

        for retry_count in range(1, 3):
            template_path = path.join(self.root_path, settings.TEMPLATE_FOLDER_NAME)
            if path.exists(template_path):
                root_path_find = True
                break
            else:
                self.root_path = path.join(self.root_path, '..')

        if not root_path_find:
            print("Error! can't find template folder")
            raise

        self.root_path = path.realpath(self.root_path)

    def set_all_path(self):
        self.result_path = path.join(self.root_path, settings.RESULT_FOLDER_NAME)
        if not path.exists(self.result_path):
            mkdir(self.result_path)

        self.setting_ini_path = path.join(self.root_path, settings.INI_FILENAME)
        self.template_path = path.join(
            path.join(self.root_path, settings.TEMPLATE_FOLDER_NAME),
            settings.TEMPLATE_FILE_NAME)

    def get_attendance_file_path(self, file_name):
        return path.join(path.join(self.root_path, settings.ATTENDANCE_FOLDER_NAME), file_name)


class GenAttendanceSummaryWeek(object):
    def __init__(self, start_attendance_path, end_attendance_path, template_path):
        super(GenAttendanceSummaryWeek, self).__init__()
        self.start_attendance_path = start_attendance_path
        self.end_attendance_path = end_attendance_path
        self.template_path = template_path
        self.is_over_month = False
        self.user_start_datetime = None
        self.user_end_datetime = None

    def start(self, user_start_datetime, user_end_datetime):
        self.user_start_datetime = user_start_datetime
        self.user_end_datetime = user_end_datetime

        if user_start_datetime.month == user_end_datetime.month:
            self.is_over_month = False
        else:
            self.is_over_month = True

        all_attendance_list = self.read_all_attendance()
        return self.fill_in_all_attendance_to_summary(all_attendance_list)

    def fill_in_all_attendance_to_summary(self, all_attendance_list):
        summary_wb = load_workbook(self.template_path)
        for sheet_index in range(0, 2):  # only use sheet0 and sheet1
            sheet = summary_wb.worksheets[sheet_index]
            self.fill_in_attendance_for_echo_name(sheet, all_attendance_list)

        return summary_wb

    def fill_in_attendance_for_echo_name(self, sheet, all_attendance_list):
        name_row, name_col = self.find_name_cell_index(sheet)
        row_count = 0
        for cell in sheet[get_column_letter(name_col)]:
            row_count += 1
            if row_count <= name_row:
                continue
            if cell.value:
                self.fill_in_attendance_one_person(sheet, cell, all_attendance_list)

    def fill_in_attendance_one_person(self, sheet, cell, all_attendance_list):
        attendance_one_person_list = self.get_attendance_by_name(all_attendance_list, cell.value)
        if not attendance_one_person_list:
            return

        attendance_datetime_list = utils.get_all_day(self.user_start_datetime, self.user_end_datetime)

        pre_month = attendance_datetime_list[0].month
        attendance_list_index = 0
        row_count = cell.row - 1
        for date in attendance_datetime_list:
            row_count += 1
            if date.month != pre_month:
                attendance_list_index = 1
            pre_month = date.month

            attendance_day_list = attendance_one_person_list[attendance_list_index][date.day]
            if not attendance_day_list:
                continue
            col_count = cell.column
            for attendance_time in attendance_day_list:
                col_count += 1
                sheet.cell(row_count, col_count, attendance_time)

    @staticmethod
    def get_attendance_by_name(all_attendance_list, name_str):
        attendance_one_person_list = []

        for attendance_one_mon in all_attendance_list:
            for attendance in attendance_one_mon:
                if attendance[settings.ATTENDANCE_NAME_KEY] == name_str:
                    attendance_one_person_list.append(attendance)
        return attendance_one_person_list

    @staticmethod
    def find_name_cell_index(sheet):
        for rows in sheet.iter_rows(max_row=sheet.max_column):
            for cell in rows:
                if cell.value == settings.SUMMARY_NAME_STR:
                    return cell.row, cell.column

    def read_all_attendance(self):
        all_attendance_list = []
        get_attendance = GetAttendance()
        if not self.is_over_month:
            all_attendance_list.append(get_attendance.get_all_attendance(self.start_attendance_path))
        else:
            all_attendance_list.append(get_attendance.get_all_attendance(self.start_attendance_path))
            all_attendance_list.append(get_attendance.get_all_attendance(self.end_attendance_path))

        return all_attendance_list


class GetAttendance(object):
    def __init__(self):
        super(GetAttendance, self).__init__()

    def get_all_attendance(self, attendance_path):
        attendance_list = []

        sheet = self.get_sheet_handle(attendance_path)
        for row_index in range(sheet.nrows):
            if sheet.cell_value(row_index, 0) == settings.ATTENDANCE_ID_STR:
                attendance_dict = self.get_attendance_one_person(sheet, row_index)
                if attendance_dict[settings.ATTENDANCE_NAME_KEY]:
                    attendance_list.append(attendance_dict)

        return attendance_list

    @staticmethod
    def get_attendance_one_person(sheet, row_index):
        attendance_dict = {settings.ATTENDANCE_ID_KEY: sheet.cell_value(row_index,
                                                                        settings.ATTENDANCE_ID_INDEX),
                           settings.ATTENDANCE_NAME_KEY: sheet.cell_value(row_index,
                                                                          settings.ATTENDANCE_NAME_INDEX),
                           settings.ATTENDANCE_DEPT_KEY: sheet.cell_value(row_index,
                                                                          settings.ATTENDANCE_DEPT_INDEX)}
        for day in range(sheet.ncols):
            day_str = sheet.cell_value(row_index - 1, day)
            if day_str:
                time_list = None
                time_str = None
                if (row_index + 1) < sheet.nrows:
                    time_str = sheet.cell_value(row_index + 1, day)
                if time_str:
                    time_list = time_str.splitlines()

                attendance_dict[int(day_str)] = time_list

        return attendance_dict

    @staticmethod
    def get_sheet_handle(attendance_path):
        wb = xlrd.open_workbook(attendance_path)
        return wb.sheet_by_index(1)


class AttendanceCLI(SetupFilePath):
    def __init__(self):
        super(AttendanceCLI, self).__init__()
        self.ini_setting_dict = None
        self.user_start_datetime = None
        self.user_end_datetime = None
        self.start_attendance_file_path = None
        self.end_attendance_file_path = None
        self.init()

    def init(self):
        self.ini_setting_dict = self.get_all_ini_settings()
        self.user_start_datetime = self.cover_start_datetime_setting_to_datetime()
        self.user_end_datetime = self.cover_end_datetime_setting_to_datetime()

        self.start_attendance_file_path = self.get_attendance_file_path(self.ini_setting_dict[IniEnum.START_FILENAME])
        self.end_attendance_file_path = self.get_attendance_file_path(self.ini_setting_dict[IniEnum.END_FILENAME])

    def get_all_ini_settings(self):
        ini_handle = utils.IniControl(self.setting_ini_path)
        return ini_handle.read_section_config(IniEnum.GENERAL_SECTION)

    def cover_start_datetime_setting_to_datetime(self):
        return utils.cover_str_to_datetime(self.ini_setting_dict[IniEnum.START_YEAR] + '/' +
                                           self.ini_setting_dict[IniEnum.START_MON] + '/' +
                                           self.ini_setting_dict[IniEnum.START_DAY])

    def cover_end_datetime_setting_to_datetime(self):
        return utils.cover_str_to_datetime(self.ini_setting_dict[IniEnum.END_YEAR] + '/' +
                                           self.ini_setting_dict[IniEnum.END_MON] + '/' +
                                           self.ini_setting_dict[IniEnum.END_DAY])

    def gen_attendance_summary_one_week(self):
        attendance_summary = GenAttendanceSummaryWeek(self.start_attendance_file_path,
                                                      self.end_attendance_file_path,
                                                      self.template_path)
        summary_wb = attendance_summary.start(self.user_start_datetime, self.user_end_datetime)
        summary_wb = self.fill_in_date_to_summary(summary_wb)

        result_file_name = datetime.now().strftime("%Y%m%d%H%M%S")+'_result.xlsx'
        summary_wb.save(path.join(self.result_path, result_file_name))

    def fill_in_date_to_summary(self, summary_workbook):
        summary_wb = summary_workbook
        for sheet_index in range(0, 2):  # only use sheet0 and sheet1
            sheet = summary_wb.worksheets[sheet_index]
            self.fill_in_date_range_str(sheet)
            self.fill_in_date_str(sheet)

        return summary_wb

    def fill_in_date_str(self, sheet):
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == settings.SUMMARY_DATE_STR:
                    datetime_list = utils.get_all_day(self.user_start_datetime, self.user_end_datetime)
                    row_count = cell.row
                    for date in datetime_list:
                        row_count += 1
                        date_str = str(date.month) + '/' + str(date.day)
                        sheet.cell(row_count, cell.column, date_str)
                    return

    def fill_in_date_range_str(self, sheet):
        for row in sheet.iter_rows():
            for cell in row:
                if str(cell.value).find(settings.SUMMARY_DATE_RANGE_STR) > -1:
                    date_range_str = settings.SUMMARY_DATE_RANGE_STR + \
                                     self.ini_setting_dict[IniEnum.START_YEAR] + '/' +\
                                     self.ini_setting_dict[IniEnum.START_MON] + '/' +\
                                     self.ini_setting_dict[IniEnum.START_DAY] + '~' +\
                                     self.ini_setting_dict[IniEnum.END_YEAR] + '/' +\
                                     self.ini_setting_dict[IniEnum.END_MON] + '/' +\
                                     self.ini_setting_dict[IniEnum.END_DAY]

                    sheet.cell(cell.row, cell.column, date_range_str)


if __name__ == '__main__':
    if len(sys.argv) == 1:
        gen_attendance_summary = AttendanceCLI()
        gen_attendance_summary.gen_attendance_summary_one_week()
    elif sys.argv[1] == 'ver':
        print(settings.SW_VERSION)
    else:
        print('Error! Parameter error')
